from openpyxl import *

wb=load_workbook("C:\\Users\\shashi\Desktop\\New folder (2)\\New folder\\ExcelSheet.xlsx")
sh=wb['Login']
max_row=sh.max_row
max_col=sh.max_column

# datalist=[]
# for i in range(1,max_row+1):
#     newrow=[]
#     for j in range(1,max_col+1):
#         val=sh.cell(row=i,column=j).value
#         newrow.append(val)
#     datalist.append(newrow)
# print(datalist,end=" ")


datalist=[]
for i in range(1,max_row+1):
    newrow=[]
    for j in range(1,max_col+1):
        val=sh.cell(i,j).value
        newrow.append(val)
    datalist.append(newrow)
print(datalist,end=" ")